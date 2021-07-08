import openpyxl
import jieba

wb1 = openpyxl.load_workbook("hospital_data/Generated - Shanghai Hospital List.xlsx")
# wb1 = openpyxl.load_workbook("hospital_data/Generated - Chengdu Hospital List.xlsx")
print(wb1.sheetnames)
sheet1 = wb1["Sheet1"]
Galderma_hospital_B_columns = sheet1["B"]

wb2 = openpyxl.load_workbook("hospital_data/Data Source - SoYoung List.xlsx")
print(wb2.sheetnames)
sheet2 = wb2["Sheet1"]
SoYoung_B_columns = sheet2["B"]

wb3 = openpyxl.load_workbook("hospital_data/Data Source - Allergan List.xlsx")
print(wb3.sheetnames)
sheet3 = wb3["Sheet1"]
Allergan_B_columns = sheet3["B"]

wb4 = openpyxl.load_workbook("hospital_data/Data Source - SAIC List.xlsx")
print(wb4.sheetnames)
sheet4 = wb4["Sheet1"]
SAIC_f_columns = sheet4["F"]

# City:
city_shanghai = "上海"
city_chengdu = "成都"

# Filter：
filer_list = ["上海", "上海市", "闵行", "虹桥", "金山", "杭州", "四川", "四川省", "西藏",
              "成都", "成都市", "崇州", "崇州市", "岳阳", "岳阳市", "浆洗街",
              "自治", "自治区", "温江", "温江区", "天府", "新区", "天府新区",
              "人民", "政府", "西藏自治区人民政府", "驻", "办事处", "科", "附属", "国际",
              "成华", "成华区", "高新", "高新区", "锦江", "锦江区", "武侯", "武侯区", "青羊", "青羊区", "金牛", "金牛区",
              "综合", "开放", "医疗", "医美", "医生", "医学", "医学美容", "医学中心",
              "医院", "美容", "美容外科", "整形外科", "眼科医院", "妇产医院", "口腔医院", "皮肤",
              "盆底", "康复", "中心", "诊所", "门诊", "门诊部", "综合", "整形", "中西医", "中医", "西医", "结合", "特色",
              "有限公司", "有限责任", "企业", "投资", "管理", "咨询", "普通", "合伙", "公司", "集团",
              "科技", "开发", "科技开发",
              "大学", "复旦", "复旦大学", "交通", "交通大学", "四川大学",
              "（", "长宁店", "）", "(", "奥克斯", "广场", "店", ")"]
print(filer_list)

Galderma_hospital_row_index = 0

for Galderma_Hospital_cell in Galderma_hospital_B_columns:
    keyword_list = []
    Galderma_hospital_row_index = Galderma_hospital_row_index + 1
    # 数据清洗："None"，"-"，"VLOOKUP"
    if Galderma_Hospital_cell.value is not None and Galderma_Hospital_cell.value != "-" and "VLOOKUP" not in Galderma_Hospital_cell.value:
        print(Galderma_Hospital_cell.value)

        # *** 1st Keyword Location *** #
        word_list = jieba.lcut(Galderma_Hospital_cell.value)
        print(", ".join(word_list))

        for word in word_list:
            if word not in filer_list:
                keyword_list.append(word)

        keyword = "".join(keyword_list)
        print("Keyword：" + keyword)

        # *** 1st SoYoung Mapping *** #
        keyword_search_SoYoung_row_index = 0
        keyword_search_SoYoung_cell_value_list = []

        for SoYoung_cell in SoYoung_B_columns:
            keyword_search_SoYoung_row_index = keyword_search_SoYoung_row_index + 1
            # 数据清洗："None"，"-"，"VLOOKUP"
            if SoYoung_cell.value is not None and SoYoung_cell.value != "-" and "VLOOKUP" not in SoYoung_cell.value:
                if city_shanghai in str(sheet2['E' + str(keyword_search_SoYoung_row_index)].value):
                    if keyword != "" and keyword in SoYoung_cell.value:
                        print("SoYoung keyword mapping row index: " + str(keyword_search_SoYoung_row_index))
                        print("SoYoung keyword mapping cell value：" + str(SoYoung_cell.value))
                        keyword_search_SoYoung_cell_value_list.append(SoYoung_cell.value)

        keyword_SoYoung_mapping_list = keyword_search_SoYoung_cell_value_list
        keyword_SoYoung_mapping_string = ", ".join(keyword_SoYoung_mapping_list)

        print("1st SoYoung Mapping: " + keyword_SoYoung_mapping_string)
        print()

        # *** 2nd SoYoung Mapping *** #
        if keyword_SoYoung_mapping_string == "" and len(keyword) > 3:
            sub_keyword_list = []

            for i in range(0, len(keyword), 2):
                sub_keyword_list.append(keyword[i:i + 2])

            print("sub keyword list:" + str(sub_keyword_list))

            for sub_keyword in sub_keyword_list:
                if len(sub_keyword) < 2:
                    break
                print("SoYoung sub Keyword：" + sub_keyword)

                sub_keyword_search_SoYoung_row_index = 0
                sub_keyword_search_SoYoung_cell_value_list = []

                if sub_keyword not in filer_list:
                    for SoYoung_cell in SoYoung_B_columns:
                        sub_keyword_search_SoYoung_row_index = sub_keyword_search_SoYoung_row_index + 1
                        # 数据清洗："None"，"-"，"VLOOKUP"
                        if SoYoung_cell.value is not None and SoYoung_cell.value != "-" and "VLOOKUP" not in SoYoung_cell.value:
                            if city_shanghai in str(sheet2['E' + str(sub_keyword_search_SoYoung_row_index)].value):
                                if sub_keyword != "" and sub_keyword in SoYoung_cell.value:
                                    print("SoYoung sub-keyword mapping row index: " + str(
                                        sub_keyword_search_SoYoung_row_index))
                                    print("SoYoung sub-keyword mapping cell value：" + str(SoYoung_cell.value))
                                    sub_keyword_search_SoYoung_cell_value_list.append(SoYoung_cell.value)
                                    break

                if sub_keyword_search_SoYoung_cell_value_list:
                    break

            keyword_SoYoung_mapping_list = sub_keyword_search_SoYoung_cell_value_list
            keyword_SoYoung_mapping_string = ", ".join(keyword_SoYoung_mapping_list)
            print("2nd SoYoung Mapping: " + keyword_SoYoung_mapping_string)
            print()

        # *** 1st Allergan Mapping *** #
        keyword_search_Allergan_row_index = 0
        keyword_search_Allergan_cell_value_list = []

        for cell in Allergan_B_columns:
            keyword_search_Allergan_row_index = keyword_search_Allergan_row_index + 1
            # 数据清洗："None"，"-"，"VLOOKUP"
            if cell.value is not None and cell.value != "-" and "VLOOKUP" not in cell.value:
                if city_shanghai in str(sheet3['M' + str(keyword_search_Allergan_row_index)].value):
                    if keyword != "" and keyword in cell.value:
                        print("Allergan keyword mapping row index: " + str(keyword_search_Allergan_row_index))
                        print("Allergan keyword mapping cell value：" + str(cell.value))
                        keyword_search_Allergan_cell_value_list.append(cell.value)
                        break

        keyword_Allergan_mapping_list = keyword_search_Allergan_cell_value_list
        keyword_Allergan_mapping_string = ", ".join(keyword_Allergan_mapping_list)
        print("1st Allergan Mapping: " + keyword_Allergan_mapping_string)
        print()

        # *** 2nd Allergan Mapping *** #
        if keyword_Allergan_mapping_string == "" and len(keyword) > 3:
            sub_keyword_list = []

            for i in range(0, len(keyword), 2):
                sub_keyword_list.append(keyword[i:i + 2])

            print("Allergan sub keyword list:" + str(sub_keyword_list))

            for sub_keyword in sub_keyword_list:
                if len(sub_keyword) < 2:
                    break

                sub_keyword_search_Allergan_row_index = 0
                sub_keyword_search_Allergan_cell_value_list = []

                if sub_keyword not in filer_list:
                    print("sub Keyword：" + sub_keyword)

                    for Allergan_cell in Allergan_B_columns:
                        sub_keyword_search_Allergan_row_index = sub_keyword_search_Allergan_row_index + 1
                        # 数据清洗："None"，"-"，"VLOOKUP"
                        if Allergan_cell.value is not None and Allergan_cell.value != "-" and "VLOOKUP" not in Allergan_cell.value:
                            if city_shanghai in str(sheet3['M' + str(sub_keyword_search_Allergan_row_index)].value):
                                if sub_keyword != "" and sub_keyword in Allergan_cell.value:
                                    print("Allergan sub-keyword mapping row index: " + str(sub_keyword_search_Allergan_row_index))
                                    print("Allergan sub-keyword mapping cell value：" + str(Allergan_cell.value))
                                    sub_keyword_search_Allergan_cell_value_list.append(Allergan_cell.value)
                                    break

                if sub_keyword_search_Allergan_cell_value_list:
                    break

            keyword_Allergan_mapping_list = sub_keyword_search_Allergan_cell_value_list
            keyword_Allergan_mapping_string = ", ".join(keyword_Allergan_mapping_list)
            print("2nd Allergan Mapping: " + keyword_Allergan_mapping_string)
            print()

        sheet1["H" + str(Galderma_hospital_row_index)] = keyword_SoYoung_mapping_string
        sheet1["L" + str(Galderma_hospital_row_index)] = keyword_Allergan_mapping_string
        # sheet1["M" + str(Galderma_hospital_row_index)] = keyword_SAIC_mapping_string

sheet1["H1"] = "SoYoung 2400 非工商"
sheet1["L1"] = "Allergan 3270 非工商 （Allergan有Gal没有的，可以考虑为D/new, esp. toxin), May-18"
# sheet1["M1"] = "SAIC 13K, May-13. 先vlookup 13K，再对USCC"

wb1.save("hospital_data/Generated - Shanghai Hospital List.xlsx")
# wb1.save("hospital_data/Generated - Chengdu Hospital List.xlsx")